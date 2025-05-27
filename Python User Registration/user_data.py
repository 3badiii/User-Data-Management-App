import customtkinter as ctk
import tkinter as tk
from openpyxl import Workbook, load_workbook
import os
from fpdf import FPDF
from datetime import datetime


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
filename = os.path.join(BASE_DIR, "user_data.xlsx")
pdf_path = os.path.join(BASE_DIR, "user_data.pdf")

if not os.path.exists(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["First Name", "Last Name", "Age", "Gender", "Status", "Saved Date"])
    wb.save(filename)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

BUTTON_STYLE = {
    "fg_color": "#00aa55",
    "hover_color": "#009944",
    "text_color": "white",
    "font": ("Arial", 14, "bold"),
    "corner_radius": 8,
    "height": 40,
    "width": 140
}

def load_data(filter_text=""):
    listbox.delete(0, tk.END)
    male_count = 0
    female_count = 0

    wb = load_workbook(filename)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 6:
            display = f"👤 {row[0]} {row[1]} | 🎂 Age: {row[2]} | ⚧ Gender: {row[3]} | 💼 Status: {row[4]} | 🕒 Saved: {row[5]}"
            if filter_text.lower() in display.lower():
                listbox.insert(tk.END, display)
            if row[3] == "Male":
                male_count += 1
            elif row[3] == "Female":
                female_count += 1

    label_stats.configure(text=f"👨‍🦰 Males: {male_count}    👩‍🦰 Females: {female_count}")

def save_data():
    fname = entry_fname.get()
    lname = entry_lname.get()
    age = entry_age.get()
    gender = gender_var.get()
    status = status_var.get()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if fname and lname and age and gender and status:
        wb = load_workbook(filename)
        ws = wb["Users"]
        ws.append([fname, lname, age, gender, status, now])
        wb.save(filename)

        Clear_box()
        label_status.configure(text="✅ Saved successfully", text_color="green")
        load_data()
    else:
        label_status.configure(text="❗️ Please fill all fields", text_color="red")

def Clear_box():
    entry_fname.delete(0, tk.END)
    entry_lname.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    gender_var.set("")
    status_var.set("")
    label_status.configure(text="")

def on_select(event):
    try:
        index = listbox.curselection()[0]
        data = listbox.get(index)
        parts = data.split("|")
        name = parts[0].replace("👤", "").strip().split()
        entry_fname.delete(0, tk.END)
        entry_fname.insert(0, name[0])
        entry_lname.delete(0, tk.END)
        entry_lname.insert(0, name[1])
        entry_age.delete(0, tk.END)
        entry_age.insert(0, parts[1].split(":")[1].strip())
        gender_var.set(parts[2].split(":")[1].strip())
        status_var.set(parts[3].split(":")[1].strip())
    except:
        pass

def delete_data():
    try:
        index = listbox.curselection()[0]
        wb = load_workbook(filename)
        ws = wb["Users"]
        ws.delete_rows(index + 2)
        wb.save(filename)
        label_status.configure(text="🗑️ Deleted successfully", text_color="orange")
        load_data()
    except:
        label_status.configure(text="❗️ Select an item to delete", text_color="red")

def update_data():
    try:
        index = listbox.curselection()[0]
        fname = entry_fname.get()
        lname = entry_lname.get()
        age = entry_age.get()
        gender = gender_var.get()
        status = status_var.get()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if fname and lname and age and gender and status:
            wb = load_workbook(filename)
            ws = wb["Users"]
            ws.delete_rows(index + 2)
            ws.insert_rows(index + 2)
            ws.cell(row=index + 2, column=1, value=fname)
            ws.cell(row=index + 2, column=2, value=lname)
            ws.cell(row=index + 2, column=3, value=age)
            ws.cell(row=index + 2, column=4, value=gender)
            ws.cell(row=index + 2, column=5, value=status)
            ws.cell(row=index + 2, column=6, value=now)
            wb.save(filename)
            label_status.configure(text="✏️ Updated successfully", text_color="cyan")
            load_data()
        else:
            label_status.configure(text="❗️ Fill all fields before update", text_color="red")
    except:
        label_status.configure(text="❗️ Select an item to update", text_color="red")

def open_excel():
    os.startfile(filename)

def search():
    text = search_var.get()
    load_data(text)

def export_pdf():
    wb = load_workbook(filename)
    ws = wb["Users"]
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="User Data", ln=True, align="C")
    pdf.ln(10)
    for row in ws.iter_rows(min_row=2, values_only=True):
        line = f"{row[0]} {row[1]} | Age: {row[2]} | Gender: {row[3]} | Status: {row[4]} | Saved: {row[5]}"
        pdf.cell(200, 10, txt=line, ln=True)
    pdf.output(pdf_path)
    label_status.configure(text="📄 PDF exported successfully", text_color="lightblue")



def login():
    user = username_entry.get()
    pwd = password_entry.get()
    if user == "admin" and pwd == "1234":
        login_win.destroy()
        launch_main_app()
    else:
        login_error.configure(text="❌ Incorrect username or password")

login_win = ctk.CTk()
login_win.title("Login")
login_win.geometry("400x300")
ctk.CTkLabel(login_win, text="🔐 Login", font=("Arial", 24, "bold"), text_color="#00FFAA").pack(pady=20)
username_entry = ctk.CTkEntry(login_win, placeholder_text="Username")
username_entry.pack(pady=10)
password_entry = ctk.CTkEntry(login_win, placeholder_text="Password", show="*")
password_entry.pack(pady=10)
ctk.CTkButton(login_win, text="Login", command=login, **BUTTON_STYLE).pack(pady=10)
login_error = ctk.CTkLabel(login_win, text="", text_color="red")
login_error.pack()



def launch_main_app():
    global win, entry_fname, entry_lname, entry_age, gender_var, status_var, label_status, label_stats, listbox, search_var
    win = ctk.CTk()
    win.title("🌟 User Management System")
    win.geometry("1200x800")
    win.minsize(900, 600)

    container = ctk.CTkFrame(win, corner_radius=20)
    container.pack(fill="both", expand=True, padx=20, pady=20)
    container.grid_columnconfigure(0, weight=1)

    form_wrapper = ctk.CTkFrame(container, corner_radius=15)
    form_wrapper.pack(anchor="center", pady=20)

    ctk.CTkLabel(form_wrapper, text="👤 User Registration", font=("Arial", 28, "bold"), text_color="#00FFAA").grid(row=0, column=0, columnspan=2, pady=10)

    label_font = ("Arial", 16)
    input_font = ("Arial", 14)

    ctk.CTkLabel(form_wrapper, text="First Name", font=label_font).grid(row=1, column=0, pady=5, sticky="e")
    entry_fname = tk.Entry(form_wrapper, font=input_font, width=30)
    entry_fname.grid(row=1, column=1, pady=5)

    ctk.CTkLabel(form_wrapper, text="Last Name", font=label_font).grid(row=2, column=0, pady=5, sticky="e")
    entry_lname = tk.Entry(form_wrapper, font=input_font, width=30)
    entry_lname.grid(row=2, column=1, pady=5)

    ctk.CTkLabel(form_wrapper, text="Age", font=label_font).grid(row=3, column=0, pady=5, sticky="e")
    entry_age = tk.Entry(form_wrapper, font=input_font, width=30)
    entry_age.grid(row=3, column=1, pady=5)

    ctk.CTkLabel(form_wrapper, text="Gender", font=label_font).grid(row=4, column=0, pady=5, sticky="e")
    gender_var = tk.StringVar()
    ctk.CTkRadioButton(form_wrapper, text="Male", variable=gender_var, value="Male").grid(row=4, column=1, sticky="w")
    ctk.CTkRadioButton(form_wrapper, text="Female", variable=gender_var, value="Female").grid(row=4, column=1, sticky="e")

    ctk.CTkLabel(form_wrapper, text="Status", font=label_font).grid(row=5, column=0, pady=5, sticky="e")
    status_var = tk.StringVar()
    status_menu = ctk.CTkOptionMenu(form_wrapper, variable=status_var, values=["Student", "Employee", "Unemployed"])
    status_menu.grid(row=5, column=1, pady=5)

    button_frame = ctk.CTkFrame(container, corner_radius=10)
    button_frame.pack(pady=15)

    ctk.CTkButton(button_frame, text="💾 Save", command=save_data, **BUTTON_STYLE).grid(row=0, column=0, padx=5)
    ctk.CTkButton(button_frame, text="🔄 Update", command=update_data, **BUTTON_STYLE).grid(row=0, column=1, padx=5)
    ctk.CTkButton(button_frame, text="🗑️ Delete", command=delete_data, **BUTTON_STYLE).grid(row=0, column=2, padx=5)
    ctk.CTkButton(button_frame, text="📂 Open Excel", command=open_excel, **BUTTON_STYLE).grid(row=1, column=0, padx=5, pady=5)
    ctk.CTkButton(button_frame, text="📄 Export PDF", command=export_pdf, **BUTTON_STYLE).grid(row=1, column=1, padx=5, pady=5)
    ctk.CTkButton(button_frame, text="🧹 Clear", command=Clear_box, **BUTTON_STYLE).grid(row=1, column=2, padx=5, pady=5)

    search_wrapper = ctk.CTkFrame(container)
    search_wrapper.pack(pady=10)

    search_var = tk.StringVar()
    search_entry = tk.Entry(search_wrapper, textvariable=search_var, font=("Arial", 14), width=50)
    search_entry.grid(row=0, column=0, padx=5)
    ctk.CTkButton(search_wrapper, text="🔍 Search", command=search, **BUTTON_STYLE).grid(row=0, column=1)

    listbox = tk.Listbox(container, height=10, bg="#222", fg="#0f0", font=("Courier New", 13))
    listbox.pack(fill="x", padx=40, pady=10)
    listbox.bind("<<ListboxSelect>>", on_select)

    label_status = ctk.CTkLabel(container, text="", font=("Arial", 14))
    label_status.pack(pady=5)

    label_stats = ctk.CTkLabel(container, text="", font=("Arial", 16))
    label_stats.pack(pady=5)

    load_data()
    win.mainloop()

login_win.mainloop()
